import sys
import unittest
import uuid
from pathlib import Path

from openpyxl import load_workbook


JULY_DIR = Path(__file__).resolve().parents[1]
WORKSPACE_TMP = JULY_DIR.parents[2] / ".tmp"
sys.path.insert(0, str(JULY_DIR))

from ado_guideline_export import DevOpsSnapshot, export_guideline, read_devops_configuration  # noqa: E402


class GuidelineExportTests(unittest.TestCase):
    def _snapshot(self) -> DevOpsSnapshot:
        return DevOpsSnapshot(
            organization_url="https://dev.azure.com/contoso",
            project={"id": "project-id", "name": "Contoso ERP"},
            process={"typeId": "process-id", "name": "Contoso Process"},
            work_item_types=[
                {
                    "name": "Requirement",
                    "referenceName": "Contoso.Requirement",
                    "customization": "custom",
                    "color": "009CCC",
                    "icon": "icon_book",
                    "description": "A project requirement.",
                }
            ],
            organization_fields=[
                {
                    "name": "Title",
                    "referenceName": "System.Title",
                    "type": "String",
                    "description": "The work item title.",
                },
                {
                    "name": "Risk rating",
                    "referenceName": "Contoso.RiskRating",
                    "type": "String",
                    "description": "Current risk rating.",
                    "isPicklist": True,
                    "picklistId": "risk-list",
                },
                {"name": "ID", "referenceName": "System.Id", "type": "Integer"},
            ],
            picklists=[
                {"id": "risk-list", "name": "Risk rating", "type": "String", "items": ["Low", "Medium", "High"]}
            ],
            behaviors=[
                {
                    "name": "Requirements",
                    "referenceName": "System.RequirementBacklogBehavior",
                    "color": "009CCC",
                    "rank": 10,
                }
            ],
            areas={
                "name": "Contoso ERP",
                "children": [
                    {"name": "Finance", "children": [{"name": "Accounts payable", "children": []}]}
                ],
            },
            iterations={
                "name": "Contoso ERP",
                "children": [
                    {"name": "Release 1", "children": [{"name": "Sprint 1", "children": []}]}
                ],
            },
            teams=[{"id": "team-id", "name": "Finance Team"}],
            wit_fields={
                "Contoso.Requirement": [
                    {"referenceName": "System.Title", "required": True},
                    {"referenceName": "Contoso.RiskRating", "allowedValues": ["Low", "Medium", "High"]},
                    {"referenceName": "System.Id"},
                ]
            },
            layouts={
                "Contoso.Requirement": {
                    "pages": [
                        {
                            "label": "Details",
                            "visible": True,
                            "sections": [
                                {
                                    "id": "Section1",
                                    "groups": [
                                        {
                                            "label": "Planning",
                                            "order": 1,
                                            "controls": [
                                                {"id": "System.Title", "label": "Title", "order": 1, "visible": True},
                                                {
                                                    "id": "Contoso.RiskRating",
                                                    "label": "Risk rating",
                                                    "order": 2,
                                                    "visible": True,
                                                },
                                            ],
                                        }
                                    ],
                                }
                            ],
                        }
                    ]
                }
            },
            wit_behaviors={
                "Contoso.Requirement": [
                    {"behavior": {"id": "System.RequirementBacklogBehavior"}, "isDefault": True}
                ]
            },
            team_settings={
                "team-id": {
                    "bugsBehavior": "asRequirements",
                    "defaultIterationMacro": "@CurrentIteration",
                    "backlogIteration": {"path": "Contoso ERP"},
                }
            },
            team_field_values={
                "team-id": {
                    "defaultValue": "Contoso ERP\\Finance",
                    "values": [{"value": "Contoso ERP\\Finance", "includeChildren": True}],
                }
            },
            generated_at="2026-07-22T12:00:00+00:00",
        )

    def test_exports_july_compatible_workbook(self):
        template = JULY_DIR / "ADO template guideline (July).xlsx"
        WORKSPACE_TMP.mkdir(exist_ok=True)
        output = WORKSPACE_TMP / f"guideline-{uuid.uuid4().hex}.xlsx"
        try:
            export_guideline(self._snapshot(), template, output)

            workbook = load_workbook(output, read_only=False, data_only=True)
            self.assertTrue(
                {"Work item types", "Fields", "Picklists", "Iteration paths", "Area paths", "Teams", "Backlogs"}
                .issubset(workbook.sheetnames)
            )
            self.assertEqual(workbook["_Export metadata"].sheet_state, "hidden")

            fields = list(workbook["Fields"].iter_rows(values_only=True))
            field_header = list(fields[0])
            field_records = [dict(zip(field_header, row)) for row in fields[1:]]
            risk = next(row for row in field_records if row["Reference name"] == "Contoso.RiskRating")
            self.assertEqual(risk["Field type"], "PicklistString")
            self.assertEqual(risk["Group name"], "Planning")
            self.assertNotIn("System.Id", {row["Reference name"] for row in field_records})

            wits = list(workbook["Work item types"].iter_rows(values_only=True))
            wit_header = list(wits[0])
            requirement = dict(zip(wit_header, wits[1]))
            self.assertEqual(requirement["Work item type"], "Requirement")
            self.assertEqual(requirement["Backlog name"], "Requirements")
            self.assertEqual(requirement["Risk rating"], "X")

            picklists = list(workbook["Picklists"].iter_rows(values_only=True))
            risk_column = list(picklists[0]).index("Risk rating")
            self.assertEqual([row[risk_column] for row in picklists[1:4]], ["Low", "Medium", "High"])

            areas = list(workbook["Area paths"].iter_rows(values_only=True))
            self.assertIn((None, "Finance", None, None, None, "Finance Team"), areas)
            iterations = list(workbook["Iteration paths"].iter_rows(values_only=True))
            self.assertIn(("Release 1", None), iterations)
            self.assertIn((None, "Sprint 1"), iterations)

            self.assertEqual(workbook["Work item types"].tables["Table24"].ref.split(":")[0], "A1")
            workbook.close()
        finally:
            output.unlink(missing_ok=True)

    def test_does_not_overwrite_template(self):
        template = JULY_DIR / "ADO template guideline (July).xlsx"
        with self.assertRaises(ValueError):
            export_guideline(self._snapshot(), template, template)

    def test_reads_project_process_and_configuration(self):
        class FakeClient:
            organization_url = "https://dev.azure.com/contoso"

            def __init__(self):
                self.paths = []

            def get_json(self, path, **kwargs):
                self.paths.append(path)
                if path.startswith("_apis/projects/"):
                    return {
                        "id": "project-id",
                        "name": "Contoso ERP",
                        "capabilities": {"processTemplate": {"templateTypeId": "default-process"}},
                    }
                if path.endswith("/classificationnodes/Areas") or path.endswith("/classificationnodes/Iterations"):
                    return {"name": "Contoso ERP", "children": []}
                return {}

            def get_collection(self, path, **kwargs):
                self.paths.append(path)
                if path == "_apis/work/processes":
                    return [
                        {"typeId": "default-process", "name": "Default"},
                        {"typeId": "override-process", "name": "Override"},
                    ]
                if path.endswith("/workitemtypes"):
                    return [{"name": "Requirement", "referenceName": "Contoso.Requirement"}]
                if path == "_apis/wit/fields":
                    return []
                if path == "_apis/work/processes/lists":
                    return []
                if path.endswith("/behaviors"):
                    return []
                if path.endswith("/teams"):
                    return []
                if "/fields" in path:
                    return []
                return []

        client = FakeClient()
        snapshot = read_devops_configuration(client, "Contoso ERP", process_name="Override")
        self.assertEqual(snapshot.process["typeId"], "override-process")
        self.assertTrue(any("override-process" in path and path.endswith("/workitemtypes") for path in client.paths))


if __name__ == "__main__":
    unittest.main()
