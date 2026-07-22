"""Export an Azure DevOps project/process to a July-compatible guideline workbook."""

from __future__ import annotations

import copy
import datetime as dt
import os
import re
import urllib.parse
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


API_VERSION = "7.1"
FIXED_WIT_COLUMNS = [
    "Work item type",
    "Custom work item type",
    "Inherit from",
    "Reference name",
    "Color",
    "Icon",
    "Help text",
    "Purpose",
    "Backlog type",
    "Backlog name",
]
FIELDS_COLUMNS = [
    "Page name",
    "Group location",
    "Group sequence",
    "Group name",
    "Field sequence",
    "Field name",
    "Label",
    "Reference name",
    "Description",
    "Field type",
    "Multi-select field",
    "Required",
    "Default value",
    "Custom field",
    "Field use recommendation",
    "Rule name",
    "When",
    "When Field",
    "When Value",
    "Then",
    "Then Field",
]


def normalize_org_url(value: str) -> str:
    value = value.strip().rstrip("/")
    if not value:
        raise ValueError("Azure DevOps organization URL is required.")
    if value.startswith("https://dev.azure.com/"):
        return value
    if value.startswith("http://") or value.startswith("https://"):
        return value
    return f"https://dev.azure.com/{value}"


def _quote(value: str) -> str:
    return urllib.parse.quote(str(value), safe="")


class AzureDevOpsReadClient:
    """Small read-only Azure DevOps REST client used by the exporter."""

    def __init__(self, organization_url: str, pat: str, *, timeout: int = 30, session: Any = None):
        self.organization_url = normalize_org_url(organization_url)
        self.timeout = timeout
        self.session = session or requests.Session()
        if session is None:
            retry = Retry(
                total=4,
                connect=4,
                read=4,
                backoff_factor=0.75,
                status_forcelist=(408, 429, 500, 502, 503, 504),
                allowed_methods=frozenset({"GET"}),
                respect_retry_after_header=True,
            )
            adapter = HTTPAdapter(max_retries=retry)
            self.session.mount("https://", adapter)
            self.session.mount("http://", adapter)
        self.session.auth = ("", pat)
        self.session.headers.update({"Accept": "application/json"})

    def get_json(
        self,
        path: str,
        *,
        params: Optional[Mapping[str, Any]] = None,
        optional_statuses: Sequence[int] = (),
    ) -> dict:
        url = path if path.startswith("http") else f"{self.organization_url}/{path.lstrip('/')}"
        response = self.session.get(url, params=dict(params or {}), timeout=self.timeout)
        if response.status_code in optional_statuses:
            return {}
        if response.status_code >= 400:
            detail = response.text[:1000]
            raise RuntimeError(f"Azure DevOps GET failed ({response.status_code}) for {url}: {detail}")
        return response.json()

    def get_collection(
        self,
        path: str,
        *,
        params: Optional[Mapping[str, Any]] = None,
        optional_statuses: Sequence[int] = (),
    ) -> List[dict]:
        # Most process endpoints are not paged. Core endpoints use $top/$skip,
        # so callers request a high enough page size for a single project.
        payload = self.get_json(path, params=params, optional_statuses=optional_statuses)
        return list(payload.get("value", [])) if payload else []


@dataclass
class DevOpsSnapshot:
    organization_url: str
    project: dict
    process: dict
    work_item_types: List[dict]
    organization_fields: List[dict]
    picklists: List[dict]
    behaviors: List[dict]
    areas: dict
    iterations: dict
    teams: List[dict]
    wit_fields: Dict[str, List[dict]] = field(default_factory=dict)
    layouts: Dict[str, dict] = field(default_factory=dict)
    wit_behaviors: Dict[str, List[dict]] = field(default_factory=dict)
    team_settings: Dict[str, dict] = field(default_factory=dict)
    team_field_values: Dict[str, dict] = field(default_factory=dict)
    generated_at: str = field(default_factory=lambda: dt.datetime.now(dt.timezone.utc).isoformat())


def read_devops_configuration(
    client: AzureDevOpsReadClient,
    project_name: str,
    *,
    process_name: str = "",
) -> DevOpsSnapshot:
    """Read all Azure DevOps configuration needed by the July workbook."""

    project = client.get_json(
        f"_apis/projects/{_quote(project_name)}",
        params={"includeCapabilities": True, "api-version": API_VERSION},
    )
    processes = client.get_collection("_apis/work/processes", params={"api-version": API_VERSION})
    capability = project.get("capabilities", {}).get("processTemplate", {})
    process_id = str(capability.get("templateTypeId") or "")
    if process_name:
        process = next(
            (item for item in processes if str(item.get("name", "")).casefold() == process_name.casefold()),
            None,
        )
    else:
        process = next((item for item in processes if process_id and str(item.get("typeId")) == process_id), None)
    if not process:
        expected = process_name or capability.get("templateName") or process_id or "the project's process"
        raise RuntimeError(f"Could not resolve Azure DevOps process '{expected}'.")
    process_id = str(process.get("typeId"))

    work_item_types = client.get_collection(
        f"_apis/work/processes/{_quote(process_id)}/workitemtypes",
        params={"api-version": API_VERSION},
    )
    organization_fields = client.get_collection("_apis/wit/fields", params={"api-version": API_VERSION})
    picklist_metadata = client.get_collection("_apis/work/processes/lists", params={"api-version": API_VERSION})
    picklists = []
    for item in picklist_metadata:
        item_id = item.get("id")
        details = client.get_json(
            f"_apis/work/processes/lists/{_quote(item_id)}",
            params={"api-version": API_VERSION},
            optional_statuses=(404,),
        ) if item_id else {}
        picklists.append(details or item)

    behaviors = client.get_collection(
        f"_apis/work/processes/{_quote(process_id)}/behaviors",
        params={"api-version": API_VERSION},
    )
    encoded_project = _quote(project.get("id") or project_name)
    areas = client.get_json(
        f"{encoded_project}/_apis/wit/classificationnodes/Areas",
        params={"$depth": 20, "api-version": API_VERSION},
    )
    iterations = client.get_json(
        f"{encoded_project}/_apis/wit/classificationnodes/Iterations",
        params={"$depth": 20, "api-version": API_VERSION},
    )
    teams = client.get_collection(
        f"_apis/projects/{encoded_project}/teams",
        params={"$top": 1000, "api-version": f"{API_VERSION}-preview.3"},
    )

    snapshot = DevOpsSnapshot(
        organization_url=client.organization_url,
        project=project,
        process=process,
        work_item_types=work_item_types,
        organization_fields=organization_fields,
        picklists=picklists,
        behaviors=behaviors,
        areas=areas,
        iterations=iterations,
        teams=teams,
    )

    for wit in work_item_types:
        wit_ref = str(wit.get("referenceName") or "")
        wit_name = str(wit.get("name") or wit_ref)
        if not wit_ref:
            continue
        snapshot.layouts[wit_ref] = client.get_json(
            f"_apis/work/processes/{_quote(process_id)}/workItemTypes/{_quote(wit_ref)}/layout",
            params={"api-version": f"{API_VERSION}-preview.1"},
            optional_statuses=(400, 404),
        )
        snapshot.wit_behaviors[wit_ref] = client.get_collection(
            f"_apis/work/processes/{_quote(process_id)}/workitemtypesbehaviors/{_quote(wit_ref)}/behaviors",
            params={"api-version": f"{API_VERSION}-preview.1"},
            optional_statuses=(400, 404),
        )
        snapshot.wit_fields[wit_ref] = client.get_collection(
            f"{encoded_project}/_apis/wit/workitemtypes/{_quote(wit_name)}/fields",
            params={"$expand": "All", "api-version": API_VERSION},
            optional_statuses=(404,),
        )

    for team in teams:
        team_key = str(team.get("id") or team.get("name") or "")
        team_route = _quote(team.get("id") or team.get("name") or "")
        if not team_key:
            continue
        snapshot.team_settings[team_key] = client.get_json(
            f"{encoded_project}/{team_route}/_apis/work/teamsettings",
            params={"api-version": API_VERSION},
            optional_statuses=(404,),
        )
        snapshot.team_field_values[team_key] = client.get_json(
            f"{encoded_project}/{team_route}/_apis/work/teamsettings/teamfieldvalues",
            params={"api-version": API_VERSION},
            optional_statuses=(404,),
        )
    return snapshot


def _headers_and_rows(sheet: Any) -> Tuple[List[str], List[dict]]:
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]})
    return headers, rows


def _contribution_field(control: Mapping[str, Any]) -> str:
    contribution = control.get("contribution") or {}
    inputs = contribution.get("inputs") or contribution.get("properties") or {}
    if isinstance(inputs, Mapping):
        for key in ("FieldName", "fieldName", "FieldReferenceName", "fieldReferenceName"):
            if inputs.get(key):
                return str(inputs[key])
    if isinstance(inputs, list):
        for item in inputs:
            if str(item.get("id", "")).casefold() in {"fieldname", "fieldreferencename"}:
                return str(item.get("value") or "")
    return ""


def _layout_locations(layout: Mapping[str, Any]) -> Dict[str, dict]:
    locations: Dict[str, dict] = {}
    for page_index, page in enumerate(layout.get("pages") or []):
        if page.get("visible") is False:
            continue
        page_name = str(page.get("label") or page.get("id") or "Details")
        for section_index, section in enumerate(page.get("sections") or []):
            section_name = str(section.get("id") or f"Section{section_index + 1}")
            for group_index, group in enumerate(section.get("groups") or []):
                if group.get("visible") is False:
                    continue
                group_name = str(group.get("label") or group.get("id") or "Additional fields")
                for control_index, control in enumerate(group.get("controls") or []):
                    if control.get("visible") is False:
                        continue
                    reference = _contribution_field(control) or str(control.get("id") or "")
                    if not reference or reference.startswith("Control."):
                        continue
                    locations.setdefault(
                        reference,
                        {
                            "Page name": page_name,
                            "Group location": section_name,
                            "Group sequence": group.get("order", group_index + 1),
                            "Group name": group_name,
                            "Field sequence": control.get("order", control_index + 1),
                            "Label": control.get("label"),
                        },
                    )
    return locations


def _field_type(field_info: Mapping[str, Any]) -> str:
    raw_type = str(field_info.get("type") or "String")
    if field_info.get("isPicklist") or field_info.get("picklistId"):
        return "PicklistInteger" if raw_type.casefold() == "integer" else "PicklistString"
    names = {
        "string": "String",
        "plainText": "PlainText",
        "html": "HTML",
        "integer": "Integer",
        "double": "Double",
        "dateTime": "DateTime",
        "treePath": "TreePath",
        "history": "History",
        "boolean": "Boolean",
        "identity": "Identity",
        "guid": "Guid",
    }
    return names.get(raw_type, names.get(raw_type.casefold(), raw_type))


def _behavior_id(item: Mapping[str, Any]) -> str:
    behavior = item.get("behavior") or {}
    return str(behavior.get("id") or behavior.get("referenceName") or item.get("id") or "")


def _behavior_type(behavior: Mapping[str, Any]) -> str:
    reference = str(behavior.get("referenceName") or "")
    if reference == "System.TaskBacklogBehavior":
        return "Iteration backlog"
    if reference == "System.RequirementBacklogBehavior":
        return "Requirements backlog"
    return "Portfolio backlog"


def _classification_paths(root: Mapping[str, Any]) -> List[List[str]]:
    paths: List[List[str]] = []

    def visit(node: Mapping[str, Any], parents: List[str]) -> None:
        current = parents + ([str(node.get("name"))] if node.get("name") else [])
        if parents and node.get("name"):
            paths.append(current[1:])  # exclude the project root
        for child in node.get("children") or []:
            visit(child, current)

    visit(root, [])
    return paths


def _relative_area(value: Any, project_name: str) -> str:
    text = str(value or "").replace("/", "\\").strip("\\")
    prefix = f"{project_name}\\"
    if text.casefold().startswith(prefix.casefold()):
        text = text[len(prefix):]
    if text.casefold() == project_name.casefold():
        return ""
    return text


def _unique_label(preferred: str, reference: str, used: set[str]) -> str:
    base = preferred.strip() or reference.rsplit(".", 1)[-1] or "Field"
    candidate = base
    counter = 2
    while candidate.casefold() in used:
        candidate = f"{base} ({counter})"
        counter += 1
    used.add(candidate.casefold())
    return candidate


def _build_export_rows(snapshot: DevOpsSnapshot, workbook: Any) -> Dict[str, Tuple[List[str], List[List[Any]]]]:
    _, existing_field_rows = _headers_and_rows(workbook["Fields"])
    existing_fields = {str(row.get("Reference name") or ""): row for row in existing_field_rows}
    _, existing_wit_rows = _headers_and_rows(workbook["Work item types"])
    existing_wits = {str(row.get("Reference name") or ""): row for row in existing_wit_rows}

    layout_by_ref: Dict[str, dict] = {}
    for layout in snapshot.layouts.values():
        for reference, location in _layout_locations(layout).items():
            layout_by_ref.setdefault(reference, location)

    org_fields = {str(item.get("referenceName") or ""): item for item in snapshot.organization_fields}
    assigned_fields: Dict[str, List[dict]] = {}
    for fields in snapshot.wit_fields.values():
        for item in fields:
            reference = str(item.get("referenceName") or "")
            if reference:
                assigned_fields.setdefault(reference, []).append(item)

    field_rows_by_ref: Dict[str, dict] = {}
    used_labels: set[str] = set()
    for reference in sorted(assigned_fields, key=lambda value: (org_fields.get(value, {}).get("name", value).casefold())):
        # The WIT fields API also returns computed/internal fields such as IDs
        # and link counts. They cannot be safely recreated by the July setup
        # scripts, so export only known guideline fields, visible layout fields,
        # and custom fields owned by the process/organization.
        is_system_field = reference.startswith(("System.", "Microsoft."))
        if is_system_field and reference not in existing_fields and reference not in layout_by_ref:
            continue
        base = dict(existing_fields.get(reference, {}))
        info = dict(org_fields.get(reference, {}))
        if not info:
            info.update(assigned_fields[reference][0])
        location = layout_by_ref.get(reference, {})
        label = _unique_label(
            str(base.get("Label") or location.get("Label") or info.get("name") or ""),
            reference,
            used_labels,
        )
        defaults = [item.get("defaultValue") for item in assigned_fields[reference] if item.get("defaultValue") is not None]
        required = any(bool(item.get("alwaysRequired") or item.get("required")) for item in assigned_fields[reference])
        row = {column: base.get(column) for column in FIELDS_COLUMNS}
        row.update(
            {
                "Page name": location.get("Page name") or base.get("Page name") or "Details",
                "Group location": location.get("Group location") or base.get("Group location") or "Section1",
                "Group sequence": (
                    location.get("Group sequence")
                    if location.get("Group sequence") is not None
                    else (base.get("Group sequence") if base.get("Group sequence") is not None else 1)
                ),
                "Group name": location.get("Group name") or base.get("Group name") or "Additional fields",
                "Field sequence": (
                    location.get("Field sequence")
                    if location.get("Field sequence") is not None
                    else (base.get("Field sequence") if base.get("Field sequence") is not None else 1)
                ),
                "Field name": info.get("name") or base.get("Field name") or label,
                "Label": label,
                "Reference name": reference,
                "Description": info.get("description") or base.get("Description") or "",
                "Field type": _field_type(info),
                "Multi-select field": (
                    "Yes"
                    if info.get("isPicklistSuggested")
                    else ("No" if info.get("isPicklist") or info.get("picklistId") else None)
                ),
                "Required": "Yes" if required else "No",
                "Default value": defaults[0] if defaults else base.get("Default value"),
                "Custom field": "No" if reference.startswith(("System.", "Microsoft.")) else "Yes",
            }
        )
        field_rows_by_ref[reference] = row
    field_rows = [field_rows_by_ref[reference] for reference in sorted(field_rows_by_ref, key=lambda ref: str(field_rows_by_ref[ref]["Label"]).casefold())]

    behaviors = {str(item.get("referenceName") or ""): item for item in snapshot.behaviors}
    wit_backlogs: Dict[str, Tuple[str, str]] = {}
    default_wits: Dict[str, str] = {}
    for wit in snapshot.work_item_types:
        wit_ref = str(wit.get("referenceName") or "")
        for association in snapshot.wit_behaviors.get(wit_ref, []):
            behavior_ref = _behavior_id(association)
            behavior = behaviors.get(behavior_ref, {})
            if behavior:
                wit_backlogs[wit_ref] = (_behavior_type(behavior), str(behavior.get("name") or ""))
                if association.get("isDefault"):
                    default_wits[behavior_ref] = str(wit.get("name") or "")
                break

    field_headers = [row["Label"] for row in field_rows]
    ref_by_label = {row["Label"]: row["Reference name"] for row in field_rows}
    wit_headers = FIXED_WIT_COLUMNS + field_headers
    wit_rows = []
    for wit in sorted(snapshot.work_item_types, key=lambda item: str(item.get("name") or "").casefold()):
        reference = str(wit.get("referenceName") or "")
        base = existing_wits.get(reference, {})
        customization = str(wit.get("customization") or "").casefold()
        custom_flag = "Disabled" if wit.get("isDisabled") else ("Yes" if customization == "custom" else "No")
        backlog_type, backlog_name = wit_backlogs.get(reference, ("Other work item types", "No associated backlog"))
        assigned = {str(item.get("referenceName") or "") for item in snapshot.wit_fields.get(reference, [])}
        values: Dict[str, Any] = {
            "Work item type": wit.get("name"),
            "Custom work item type": custom_flag,
            "Inherit from": wit.get("inherits"),
            "Reference name": reference,
            "Color": str(wit.get("color") or "").lstrip("#"),
            "Icon": wit.get("icon"),
            "Help text": wit.get("description") or base.get("Help text") or "",
            "Purpose": base.get("Purpose"),
            "Backlog type": backlog_type,
            "Backlog name": backlog_name,
        }
        values.update({label: "X" if ref_by_label[label] in assigned else None for label in field_headers})
        wit_rows.append([values.get(header) for header in wit_headers])

    picklists_by_id = {str(item.get("id") or ""): item for item in snapshot.picklists}
    picklist_columns: List[Tuple[str, List[Any]]] = []
    for row in field_rows:
        reference = str(row["Reference name"])
        info = org_fields.get(reference, {})
        values: List[Any] = []
        picklist_id = str(info.get("picklistId") or "")
        if picklist_id and picklist_id in picklists_by_id:
            values = list(picklists_by_id[picklist_id].get("items") or [])
        if not values:
            for assigned in assigned_fields.get(reference, []):
                if assigned.get("allowedValues"):
                    values = list(assigned["allowedValues"])
                    break
        if values:
            picklist_columns.append((str(row["Label"]), list(dict.fromkeys(values))))
    picklist_headers = [name for name, _ in picklist_columns]
    max_picklist_rows = max((len(values) for _, values in picklist_columns), default=0)
    picklist_rows = [
        [values[index] if index < len(values) else None for _, values in picklist_columns]
        for index in range(max_picklist_rows)
    ]

    backlog_rows = []
    for behavior in sorted(snapshot.behaviors, key=lambda item: (item.get("rank", 9999), str(item.get("name") or "").casefold())):
        behavior_ref = str(behavior.get("referenceName") or "")
        backlog_type = _behavior_type(behavior)
        rename_from = "(rename)" if backlog_type in {"Requirements backlog", "Iteration backlog"} else "(new)"
        backlog_rows.append(
            [
                behavior.get("name"),
                backlog_type,
                str(behavior.get("color") or "").lstrip("#") or None,
                default_wits.get(behavior_ref),
                rename_from,
            ]
        )

    project_name = str(snapshot.project.get("name") or "")
    teams_by_area: Dict[str, List[str]] = {}
    team_rows = []
    for team in sorted(snapshot.teams, key=lambda item: str(item.get("name") or "").casefold()):
        team_key = str(team.get("id") or team.get("name") or "")
        name = str(team.get("name") or "")
        settings = snapshot.team_settings.get(team_key, {})
        field_values = snapshot.team_field_values.get(team_key, {})
        values = field_values.get("values") or []
        for value in values:
            relative = _relative_area(value.get("value"), project_name)
            if relative:
                teams_by_area.setdefault(relative.casefold(), []).append(name)
        include_children = bool(values) and all(bool(value.get("includeChildren")) for value in values)
        backlog_iteration = settings.get("defaultIterationMacro")
        if str(backlog_iteration or "").casefold() == "@currentiteration":
            backlog_iteration = "@currentIteration"
        else:
            backlog_iteration = settings.get("backlogIteration", {}).get("path") or "@currentIteration"
        team_rows.append([name, settings.get("bugsBehavior") or "asRequirements", "Yes" if include_children else "No", backlog_iteration])

    area_rows: List[List[Any]] = [[project_name, None, None, None, None, None]]
    for path in _classification_paths(snapshot.areas):
        if len(path) > 4:
            raise ValueError(f"Area path '{'/'.join(path)}' is deeper than the four levels supported by the July package.")
        key = "\\".join(path).casefold()
        assigned_teams = teams_by_area.get(key) or [None]
        for team_name in assigned_teams:
            levels = path + [None] * (4 - len(path))
            area_rows.append([None] + levels + [team_name])

    iteration_paths = _classification_paths(snapshot.iterations)
    max_iteration_depth = max((len(path) for path in iteration_paths), default=1)
    iteration_headers = [f"Level {index}" for index in range(1, max_iteration_depth + 1)]
    iteration_rows = []
    for path in iteration_paths:
        row = [None] * max_iteration_depth
        row[len(path) - 1] = path[-1]
        iteration_rows.append(row)

    return {
        "Work item types": (wit_headers, wit_rows),
        "Fields": (FIELDS_COLUMNS, [[row.get(column) for column in FIELDS_COLUMNS] for row in field_rows]),
        "Picklists": (picklist_headers or ["No picklists"], picklist_rows or [[None]]),
        "Iteration paths": (iteration_headers, iteration_rows),
        "Area paths": (["Areas", "Level 1", "Level 2", "Level 3", "Level 4", "Teams"], area_rows),
        "Teams": (["Teams", "Bug behavior", "Include sub areas", "Backlog iteration"], team_rows),
        "Backlogs": (["Backlog name", "Backlog type", "Color", "Default work item type", "Rename from"], backlog_rows),
    }


def _copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy.copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.alignment = copy.copy(source.alignment)
    target.protection = copy.copy(source.protection)


def _replace_sheet(sheet: Any, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
    old_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    old_header_cells = {header: copy.copy(sheet.cell(1, index + 1)) for index, header in enumerate(old_headers) if header}
    old_data_cells = {header: copy.copy(sheet.cell(2, index + 1)) for index, header in enumerate(old_headers) if header}
    old_widths = {
        header: sheet.column_dimensions[get_column_letter(index + 1)].width
        for index, header in enumerate(old_headers)
        if header
    }
    fallback_header = copy.copy(sheet.cell(1, min(max(len(old_headers), 1), sheet.max_column)))
    fallback_data = copy.copy(sheet.cell(2, min(max(len(old_headers), 1), sheet.max_column)))

    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(1, column, header)
        _copy_cell_style(old_header_cells.get(header, fallback_header), cell)
        source_index = old_headers.index(header) + 1 if header in old_headers else None
        letter = get_column_letter(column)
        if source_index:
            sheet.column_dimensions[letter].width = old_widths.get(header)
        elif sheet.column_dimensions[letter].width is None:
            sheet.column_dimensions[letter].width = min(max(len(str(header)) + 2, 12), 32)
    for row_index, values in enumerate(rows, 2):
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_index, column, value)
            header = headers[column - 1]
            _copy_cell_style(old_data_cells.get(header, fallback_data), cell)

    last_row = max(2, len(rows) + 1)
    last_col = get_column_letter(max(1, len(headers)))
    for table in sheet.tables.values():
        table.ref = f"A1:{last_col}{last_row}"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{last_col}{last_row}"


def export_guideline(snapshot: DevOpsSnapshot, template_path: os.PathLike[str] | str, output_path: os.PathLike[str] | str) -> Path:
    """Build a July-compatible guideline workbook without modifying the template."""

    template = Path(template_path).resolve()
    output = Path(output_path).resolve()
    if template == output:
        raise ValueError("Output path must be different from the source template path.")
    if not template.exists():
        raise FileNotFoundError(f"Guideline template not found: {template}")

    workbook = load_workbook(template)
    required = {"Work item types", "Fields", "Picklists", "Iteration paths", "Area paths", "Teams", "Backlogs"}
    missing = sorted(required - set(workbook.sheetnames))
    if missing:
        raise ValueError("Guideline template is missing required sheets: " + ", ".join(missing))
    export_rows = _build_export_rows(snapshot, workbook)
    for sheet_name, (headers, rows) in export_rows.items():
        _replace_sheet(workbook[sheet_name], headers, rows)

    if "_Export metadata" in workbook.sheetnames:
        del workbook["_Export metadata"]
    metadata = workbook.create_sheet("_Export metadata")
    metadata.sheet_state = "hidden"
    metadata.append(["Property", "Value"])
    metadata.append(["Organization", snapshot.organization_url])
    metadata.append(["Project", snapshot.project.get("name")])
    metadata.append(["Process", snapshot.process.get("name")])
    metadata.append(["Process ID", snapshot.process.get("typeId")])
    metadata.append(["Generated (UTC)", snapshot.generated_at])
    metadata.append(["Format", "Business Process Catalog Azure DevOps setup package - July"])
    metadata.column_dimensions["A"].width = 22
    metadata.column_dimensions["B"].width = 80

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    # Reopen the file so corrupt relationship/table edits fail before handoff.
    check = load_workbook(output, read_only=True, data_only=True)
    try:
        if not required.issubset(check.sheetnames):
            raise RuntimeError("Generated guideline workbook did not retain all required sheets.")
    finally:
        check.close()
    return output
